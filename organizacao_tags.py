from mailchimp_marketing import Client
import openpyxl
import configuration as cfg
import requests
from openpyxl import Workbook, workbook
from mailchimp_marketing.api_client import ApiClientError
import json
import hashlib

chave_api = cfg.credenciais['chave_api']
servidor = cfg.credenciais['servidor']
lista = cfg.credenciais['lista']
root = 'https://'+servidor+'.api.mailchimp.com/3.0/'

parametros_membros = {
        'Authorization': 'apiKey '+chave_api
}

def obter_hash_inscrito(email):
    email = email.lower().encode()
    m = hashlib.md5(email)
    return m.hexdigest()

def tags_para_lista(membro):
        tags = []
        if(membro['tags']):
                aux=[]
                for tag in membro['tags']:
                        tags.append(tag['name'])
        else:
                tags.append('')
        retorno = ''
        for elemento in tags:
                retorno += ' '+elemento.lower()
        return traduzir_tags(retorno)

def substituir_padrao(busca,tags,retorno):
        if(tags.find(busca.lower())>0):
                retorno.append({'name': busca.upper(), 'status':'active'})
        else:
                retorno.append({'name': busca.upper(), 'status':'inactive'})
        return retorno

def traduzir_tags(tags):
        padrao = ['aluno', 'interesse','empresa','autocad','revit','sketchup','excel','design gráfico',
                'edição de vídeo','office','produto','promob','solidworks','design jogos']
        retorno=[]
        for palavra in padrao:
                retorno = substituir_padrao(palavra,tags,retorno)
        return retorno

class Obter_dados_planilha:
        def __init__(self):
                ws = openpyxl.load_workbook('./ContatosMailChimp.xlsx',data_only=True)['contatos']
                contatos = []
                for row in range(1,ws.max_row+1):
                        contato=[]
                        contato.append(str(ws.cell(row,1).value))
                        contato.append(str(ws.cell(row,2).value))
                        contato.append(str(ws.cell(row,3).value))
                        contato.append(str(ws.cell(row,4).value))
                        tags = str(ws.cell(row,5).value).split(', ')
                        if(tags == ['None']):
                                contato.append([])
                        else:
                                contato.append(tags)
                        contatos.append(contato)

class Mailchimp_Consumidor:
        def __init__(self):
                self.mailchimp = Client()
                self.mailchimp.set_config({
                "api_key": cfg.credenciais['chave_api'],
                "server": cfg.credenciais['servidor']
                })
        
        def obter_membros(self):
                resposta = requests.get(root+'lists/'+lista+'/members?count=1000&offset=0',headers=parametros_membros).json()
                membros = resposta['members']

                informacoes_membros = [
                        [
                        membro['email_address'].lower() if membro['email_address'] != '' else '',
                        membro['merge_fields']['FNAME'].lower().capitalize() if membro['merge_fields']['FNAME'] != '' else '',
                        membro['merge_fields']['LNAME'].lower().capitalize() if membro['merge_fields']['LNAME'] != '' else '',
                        membro['status'],
                        tags_para_lista(membro)
                        ] for membro in membros]
                return informacoes_membros
        
        def adicionar_membro(self,membro):
                data = {
                        'email_address':membro[0],
                        'status':membro[3],
                        'merge_fields':{
                                'FNAME':membro[1],
                                'LNAME':membro[2]
                        }}
                try:
                        resposta = requests.post(root+'lists/'+lista+'/members/',headers=parametros_membros,json=data)
                        print(resposta)
                except ApiClientError as error:
                        print("Error: {}".format(error.text))
        
        def atualizar_membro(self,membro):
                tags = {
                        'tags':[
                                {'name':'ALUNOS AUTOCAD 2D','status':'inactive'},
                                {'name':'ALUNOS DESIGN GRÁFICO','status':'inactive'},
                                {'name':'ALUNOS EXCEL AVANÇADO','status':'inactive'},
                                {'name':'ALUNOS OFFICE ESSENTIALS','status':'inactive'},
                                {'name':'ALUNOS REVIT','status':'inactive'},
                                {'name':'ALUNOS SKETCHUP','status':'inactive'},
                                {'name':'EMPRESAS','status':'inactive'},
                                {'name':'INTERESSE EDIÇÃO DE VÍDEO','status':'inactive'},
                                {'name':'INTERESSE OFFICE ESSENTIALS','status':'inactive'},
                                {'name':'INTERESSE POWER BI/EXCEL AVANÇADO','status':'inactive'},
                                {'name':'INTERESSE PRODUTOS','status':'inactive'},
                                {'name':'INTERESSE PROMOB','status':'inactive'},
                                {'name':'INTERESSE SKETCHUP','status':'inactive'},
                                {'name':'INTERESSE SOLIDWORKS','status':'inactive'},
                                {'name':'INTERESSES AUTOCAD/REVIT','status':'inactive'},
                                {'name':'INTERESSES DESIGN GRÁFICO','status':'inactive'},
                                {'name':'INTERESSES DESIGN JOGOS','status':'inactive'}
                        ]
                }
                for tag in tags['tags']:
                        membro[4].append(tag)
                tag = {'tags':membro[4]}
                try:
                        resposta = requests.post(root+'lists/'+lista+'/members/'+obter_hash_inscrito(membro[0])+'/tags',headers=parametros_membros,json=tag)
                        print(resposta)
                except ApiClientError as error:
                        print("Error: {}".format(error.text))
        
        def salvar_planilha_contatos(self,audiencia):
                wb = Workbook()
                ws = wb.create_sheet('contatos',0)
                for contato in audiencia:
                        ws.append(contato)
                wb.save('./ContatosMailChimp.xlsx')
        
        def obter_tags(self):
                resposta = requests.get(root+'lists/'+lista+'/tag-search?count=100',headers=parametros_membros).json()
                [print(respos['name']) for respos in resposta['tags']]

bot = Mailchimp_Consumidor()
[bot.atualizar_membro(x) for x in bot.obter_membros()]
#Obter_dados_planilha()
#Converter_lista()